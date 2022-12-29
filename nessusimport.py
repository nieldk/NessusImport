import time
import pandas as pd
import openpyxl
import requests

# Obtain an API access key and secret from Nessus Professional
access_key = '<YOUR_API_ACCESS_KEY>'
secret_key = '<YOUR_API_SECRET>'

# Use the requests package to make an API request to the Nessus Professional server to obtain an access token
headers = {
    'X-ApiKeys': f'accessKey={access_key}; secretKey={secret_key}'
}

response = requests.post('https://<NESSUS_SERVER_ADDRESS>/session', headers=headers)

# Check the status code of the response
if response.status_code == 200:
    # Parse the JSON response and extract the access token
    data = response.json()
    access_token = data['token']

# Use the requests package to make an API request to the Nessus Professional server to list the available scan reports
headers = {
    'X-Cookie': f'token={access_token}'
}

response = requests.get('https://<NESSUS_SERVER_ADDRESS>/scans', headers=headers)

# Check the status code of the response
if response.status_code == 200:
    # Parse the JSON response and extract the list of scan reports
    data = response.json()
    scan_reports = data['scans']

# Iterate through the list of scan reports
for scan_report in scan_reports:
    # Check if the scan is finished
    if scan_report['status'] == 'completed':
        # Download the report in CSV format
        report_id = scan_report['id']

        headers = {
            'X-Cookie': f'token={access_token}',
            'Content-Type': 'application/octet-stream'
        }

        response = requests.get(f'https://<NESSUS_SERVER_ADDRESS>/scans/{report_id}/export', headers=headers)

        # Check the status code of the response
        if response.status_code == 200:
            # Write the report data to a CSV file
            with open(f'{report_id}.csv', 'wb') as f:
                f.write(response.content)

            # Read the Nessus scan report in CSV format into a Pandas DataFrame
            df = pd.read_csv(f'{report_id}.csv')

            # Sort the DataFrame by severity
            df = df.sort_values(by='severity', ascending=False)

            # Create a new Excel workbook
            workbook = openpyxl.Workbook()

            # Add a new worksheet to the workbook
            worksheet = workbook.create_sheet()

            # Iterate through the rows and columns of the DataFrame and write the values to the worksheet
            for row in df.index:
                for col in df.columns:
                    worksheet.cell(row=row+1, column=col+1).value = df.iloc[row, col]

            # Save the work
workbook.save(f'{report_id}.xlsx')
